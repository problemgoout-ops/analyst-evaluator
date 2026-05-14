#!/usr/bin/env python3
"""
GAP-анализ компетенций аналитиков.
Сопоставляет самооценку с матрицей требований целевого грейда.

Использование:
  python3 gap_analysis.py <matrix.xlsx> <self_assessment.xlsx> [--target middle] [--name "Имя"]
"""

import openpyxl
import argparse
import sys
import json

SCORE_MAP = {
    'Не знаю': 0,
    'Знаю': 1,
    'Использую': 2,
    'Профи': 3
}

LABELS = {0: '❌ Не знаю', 1: '📖 Знаю', 2: '🛠 Использую', 3: '🦾 Профи'}
REQUIREMENTS = {1: 'Знать теорию', 2: 'Применять на практике', 3: 'Профи-уровень'}

GRADE_COLS = {
    'intern': 5,
    'junior': 6,
    'middle': 7,
    'middle_plus': 8,
    'senior': 9
}

def load_matrix(path):
    """Загружает матрицу компетенций."""
    wb = openpyxl.load_workbook(path)
    ws = wb['Перечень компетенций']
    matrix = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=9, values_only=True):
        if row[1] and row[2]:
            key = str(row[2]).strip()
            grades = {}
            for grade, col in GRADE_COLS.items():
                val = row[col - 1]
                try:
                    grades[grade] = int(val)
                except:
                    grades[grade] = 0
            matrix[key] = {
                'block': str(row[1]),
                'skill': str(row[2]),
                'desc': str(row[3]) if row[3] else '',
                'grades': grades
            }
    return matrix

def load_self_assessment(path, name=None):
    """Загружает самооценку аналитиков."""
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    headers = [str(c) if c else '' for c in next(ws.iter_rows(min_row=1, max_row=1, max_col=ws.max_column, values_only=True))]
    
    results = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        if not row[0]:
            continue
        person = str(row[0])
        if name and name not in person:
            continue
        scores = {}
        for j in range(1, len(row)):
            val = str(row[j]) if row[j] else ''
            score = 0
            for k, v in SCORE_MAP.items():
                if k in val:
                    score = v
                    break
            header = headers[j] if j < len(headers) else ''
            scores[header] = score
        results.append({'name': person, 'scores': scores})
    return results

def gap_analysis(matrix, person_scores, target='middle'):
    """Выполняет GAP-анализ для одного аналитика."""
    gaps = []
    for skill_key, skill_data in matrix.items():
        target_req = skill_data['grades'].get(target, 0)
        # Find matching score
        current = 0
        for header, score in person_scores.items():
            if skill_key in header or header.strip() in skill_key:
                current = score
                break
        
        if target_req > 0:
            gap = target_req - current
            if gap > 0:
                gaps.append({
                    'gap': gap,
                    'block': skill_data['block'],
                    'skill': skill_data['skill'],
                    'desc': skill_data['desc'],
                    'current': current,
                    'target_req': target_req,
                    'current_label': LABELS[current],
                    'target_label': REQUIREMENTS[target_req]
                })
    gaps.sort(key=lambda x: -x['gap'])
    return gaps

def main():
    parser = argparse.ArgumentParser(description='GAP-анализ компетенций аналитиков')
    parser.add_argument('matrix', help='Путь к матрице компетенций (xlsx)')
    parser.add_argument('assessment', help='Путь к файлу самооценки (xlsx)')
    parser.add_argument('--target', default='middle', choices=GRADE_COLS.keys(), help='Целевой грейд')
    parser.add_argument('--name', default=None, help='Имя аналитика (фильтр)')
    parser.add_argument('--format', default='text', choices=['text', 'json'], help='Формат вывода')
    args = parser.parse_args()

    matrix = load_matrix(args.matrix)
    people = load_self_assessment(args.assessment, args.name)
    
    if not people:
        print("Аналитики не найдены", file=sys.stderr)
        sys.exit(1)

    for person in people:
        gaps = gap_analysis(matrix, person['scores'], args.target)
        
        if args.format == 'json':
            output = {
                'name': person['name'],
                'target': args.target,
                'total_gaps': len(gaps),
                'critical': len([g for g in gaps if g['gap'] >= 2]),
                'gaps': gaps
            }
            print(json.dumps(output, ensure_ascii=False, indent=2))
        else:
            print(f"\n{'='*60}")
            print(f"GAP-анализ: {person['name']} → {args.target}")
            print(f"{'='*60}")
            print(f"Всего пробелов: {len(gaps)}")
            print(f"Критических (gap≥2): {len([g for g in gaps if g['gap'] >= 2])}")
            
            critical = [g for g in gaps if g['gap'] >= 2]
            if critical:
                print(f"\n🔴 КРИТИЧЕСКИЕ:")
                for g in critical:
                    print(f"  {g['skill']}: {g['current_label']} → {g['target_label']} (GAP={g['gap']})")
            
            minor = [g for g in gaps if g['gap'] == 1]
            if minor:
                print(f"\n🟡 НЕДОСТАТОЧНЫЕ:")
                for g in minor:
                    print(f"  {g['skill']}: {g['current_label']} → {g['target_label']}")

if __name__ == '__main__':
    main()